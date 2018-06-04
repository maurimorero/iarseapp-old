# -*- coding: utf-8 -*-

from django.shortcuts import render

# Create your views here.
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.views.generic import CreateView
from django.core.urlresolvers import reverse_lazy
from forms import RegistroForm,ProfileForm, UserForm
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import redirect
from ..survey.models import ResponseMgr
from django.contrib import messages
from ..indicadores.models import Evaluacion
from ..usuarios.models import Profile
from django.core.mail import EmailMessage, EmailMultiAlternatives
from openpyxl import Workbook
from django.http.response import HttpResponse

#class RegistroUsuario(CreateView):
#    model= User
#    template_name= "usuarios/registrar.html"
#    form_class = RegistroForm
#    second_form_class= ProfileForm
#    success_url = reverse_lazy('usuarios:usuarios_home')

def RegistroUsuario(request):
    if request.method == 'POST':
        user_form = RegistroForm(request.POST)
        profile_form = ProfileForm(request.POST)
        if user_form.is_valid() and profile_form.is_valid():
            usuario=user_form.save()
            profile_form = ProfileForm(request.POST, instance=usuario.profile)
            profile=profile_form.save()
            usuario.email=profile.email
            usuario.save()
            html_content = '<p>El usuario <strong>'+usuario.username+' - '+profile.empresa+'</strong> se ha registrado en INDICAGRO.</p>'
            email = EmailMultiAlternatives('Usuario ' + usuario.username + ' creado exitosamente','Este es el mensaje del email',
                                 to=['indicagro@bccba.org.ar'],bcc=['maurimorero@gmail.com'])
            email.attach_alternative(html_content, "text/html")
            email.send()
            print("se mando el email")
            messages.success(request, 'Ya te encuentras habilitado para aplicarte IndicAgro. '
                                      'Gracias por formar parte de un #AgroComprometido!')
            return redirect('usuarios:usuarios_home')
    else:
        user_form = RegistroForm()
        profile_form = ProfileForm()
    return render(request, 'usuarios/registrar.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

def Creditos (request):
    return render(request, 'creditos.html', {})

def Creditos1 (request):
    return render(request, 'creditos1.html', {})

def Metodologia (request):
    return render(request, 'metodologia.html', {})

def Metodologia1 (request):
    return render(request, 'metodologia1.html', {})

def UserHome(request):
    # chequea si el usuario esta habilitado
    mensaje = 0
    perfil = Profile.objects.get(user=request.user)
    if (perfil.habilitado == False):
        mensaje = 1
    pendientes= ResponseMgr.objects.filter(user=request.user,completada=False).order_by('-created')
    evaluaciones = Evaluacion.objects.filter(usuario=request.user).order_by('-fecha')
    indicadoresTodos = Evaluacion.objects.order_by('-fecha')
    perfiles = Profile.objects.all()

    if request.method == 'POST':
        # Creamos el libro de trabajo
        wb = Workbook()

        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active

        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Reporte de INDICAGRO'

        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')

        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'Empresa'
        ws['C3'] = 'CUIT'
        ws['D3'] = 'N° BPA'
        cont = 4

        informes = Evaluacion.objects.all().order_by('-fecha')

        # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for informe in informes:
            empresa=Profile.objects.filter(user=informe.usuario)
            ws.cell(row=cont, column=2).value = empresa[0].empresaCom
            ws.cell(row=cont, column=3).value = empresa[0].user.username
            ws.cell(row=cont, column=4).value = informe.id
            cont = cont + 1

        # Establecemos el nombre del archivo
        nombre_archivo = "ReporteIndicagro.xlsx"

        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

    contexto = {
        'pendientes': pendientes,
        'evaluaciones':evaluaciones,
        'indicadoresTodos':indicadoresTodos,
        'perfiles':perfiles,
        'mensaje':str(mensaje)
    }
    return render(request,'home.html',contexto)

def UserHomeErr(request,errorMsg):
    # chequea si el usuario esta habilitado
    mensaje = 0
    perfil = Profile.objects.get(user=request.user)
    if (perfil.habilitado == False):
        mensaje = 1
    pendientes= ResponseMgr.objects.filter(user=request.user,completada=False).order_by('-created')
    evaluaciones = Evaluacion.objects.filter(usuario=request.user).order_by('-fecha')
    indicadoresTodos = Evaluacion.objects.order_by('-fecha')
    perfiles = Profile.objects.all()
    contexto = {
        'pendientes': pendientes,
        'evaluaciones': evaluaciones,
        'indicadoresTodos': indicadoresTodos,
        'errorMsg':errorMsg,
        'perfiles': perfiles,
        'mensaje': str(mensaje)
    }
    return render(request,'home.html',contexto)

@login_required
@transaction.atomic
def update_profile(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            #messages.success(request, _('Your profile was successfully updated!'))
            return redirect('usuarios:usuarios_home')
        #else:
            #messages.error(request, _('Please correct the error below.'))
    else:
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=request.user.profile)
    return render(request, 'usuarios/actualizar.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

def Simulador1 (request):
    return render(request, 'simulador/1.html', {})

def Simulador2 (request):
    return render(request, 'simulador/2.html', {})

def Simulador3 (request):
    return render(request, 'simulador/3.html', {})

def Simulador4 (request):
    return render(request, 'simulador/4.html', {})

def Simulador5 (request):
    return render(request, 'simulador/5.html', {})

def Simulador6 (request):
    return render(request, 'simulador/6.html', {})

def Simulador7 (request):
    return render(request, 'simulador/7.html', {})

def Condiciones (request):
    return render(request, 'condiciones.html', {})

def Landing (request):
    return render(request, 'landing.html', {})
